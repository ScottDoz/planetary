# Licensed under a 3-clause BSD style license - see LICENSE.rst

# This sub-module is destined for common non-package specific utility
# functions.

import os
import pandas as pd

import sys
from pathlib import Path # PurePath, PurePosixPath, PureWindowsPath


#%% Directory lookups

# Data Directory function
def get_root_dir():
    '''
    This function returns the absolute path for the data directory within the
    package.

    Returns
    -------
    ROOT_DIR : Str
        Absolute path for data directory.

    '''
    
    # Get root directory from location of 'requirements.txt'
    # ROOT_DIR = os.path.dirname(os.path.abspath('requirements.txt'))
    ROOT_DIR = os.path.split(os.path.dirname(os.path.abspath('datasets.py')))[0]
    
    return ROOT_DIR

# Data Directory function
def get_data_dir():
    '''
    This function returns the absolute path for the data directory of the
    package. This can be set by the 'SR_TOOLS_DATA' environment variable.

    Returns
    -------
    DATA_DIR : Str
        Absolute path for data directory.

    '''
    
    # Look for Environment variable 'SR_TOOLS_DATA'
    envvar = os.environ.get('SR_TOOLS_DATA')
    
    if envvar is None:
        # Data path is not set in Environment variable.
        # Use a default data storage location at '~\sr_data\Data'
        DATA_DIR = Path.home()/'sr_data'/'Data'
    else:
        # Use data location set in Environment variable
        AATA_DIR = Path(envvar)
    
    
    # Check if directory exists and create
    if not os.path.exists(str(DATA_DIR)):
        os.makedirs(str(DATA_DIR))
    
    return DATA_DIR

get_data_home = get_data_dir


#%% Pandas utilities

# Pandas write to excel
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    
    https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
            

